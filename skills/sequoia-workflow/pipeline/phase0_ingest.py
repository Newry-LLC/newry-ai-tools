"""Phase 0 — Ingest + filter (deterministic, no LLM).

Parses a FINTRX market export (xlsx), computes %HNW, applies Jack's
prioritization thresholds, and writes data/firms.json.

Rules (Jack, confirmed 2026-07-01 working-session design):
  - pct_hnw = "$ High Net Worth Clients" / "Total AUM"
  - Total AUM missing or 0        -> skipped (listed in report, not in output)
  - pct_hnw < 0.60                -> discarded (listed in report, not in output)
  - 0.60 <= pct_hnw < 0.70        -> LOW
  - pct_hnw >= 0.70               -> HIGH if 3yr AUM change > 0, else MEDIUM

Usage:
  python phase0_ingest.py "<path to FINTRX xlsx>" [--out data/firms.json]
"""

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

# FINTRX column header -> our field name. Headers must match exactly;
# a missing header is a hard error so silent schema drift can't happen.
COLUMNS = {
    "Firm CRD": "firm_crd",
    "Firm Name": "name",
    "Website Address": "website",
    "Main Office City": "city",
    "Main Office State": "state",
    "Total AUM": "aum",
    "$ High Net Worth Clients": "hnw_aum",
    "# High Net Worth Clients": "hnw_clients",
    "3 Year AUM Change": "three_yr_aum_change",
    "3 Year Account Change": "three_yr_account_change",
    "Total Client Count": "total_clients",
    "Total Accounts": "accounts",
    "Employees": "employees",
    "Fee Structure": "fee_structure",
    "Average Client Size": "avg_client_size",
    "Additional Offices": "additional_offices",
    "Industry Accolades": "accolades",
    "Retail Custodian": "custodian",
    "Tamps Used": "tamps",
    "Last ADV Filing date": "last_adv_filing",
    "Firm FINTRX Profile Link": "fintrx_url",
}


def to_float(value):
    """Parse a number that may arrive as float, int, '35%', or '1,234'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    pct = s.endswith("%")
    if pct:
        s = s[:-1]
    try:
        n = float(s)
    except ValueError:
        return None
    return n / 100 if pct else n


def classify(pct_hnw, three_yr_change):
    if pct_hnw < 0.60:
        return None  # discard
    if pct_hnw < 0.70:
        return "Low"
    if three_yr_change is not None and three_yr_change > 0:
        return "High"
    return "Medium"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="FINTRX export path")
    ap.add_argument("--out", default=None, help="output JSON path")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    out = Path(args.out) if args.out else Path(__file__).parent / "data" / "firms.json"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    missing = [h for h in COLUMNS if h not in header_idx]
    if missing:
        sys.exit(f"FATAL: expected FINTRX columns not found: {missing}")

    firms, skipped, discarded = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {ours: row[header_idx[theirs]] for theirs, ours in COLUMNS.items()}
        if rec["name"] is None:
            continue  # blank row
        # JSON-safe: dates -> ISO strings
        for k, v in rec.items():
            if isinstance(v, (datetime, date)):
                rec[k] = v.isoformat()

        aum = to_float(rec["aum"])
        hnw = to_float(rec["hnw_aum"])
        if not aum:
            skipped.append(rec["name"])
            continue

        pct_hnw = (hnw or 0.0) / aum
        rec["pct_hnw"] = round(pct_hnw, 4)
        # FINTRX percent columns arrive as raw numbers ('35' meaning 35%) or
        # '35%' strings; to_float normalizes '%' strings to fractions. Bare
        # numbers > 5 are assumed to be percentage points.
        change = to_float(rec["three_yr_aum_change"])
        if change is not None and abs(change) > 5:
            change = change / 100
        rec["three_yr_aum_change_frac"] = change

        priority = classify(pct_hnw, change)
        if priority is None:
            discarded.append(f"{rec['name']} ({pct_hnw:.0%} HNW)")
            continue
        rec["priority"] = priority
        rec["firm_crd"] = str(rec["firm_crd"]).strip() if rec["firm_crd"] else None
        firms.append(rec)

    firms.sort(key=lambda r: (-r["pct_hnw"]))
    payload = {
        "source_file": str(xlsx),
        "market": "Phoenix",
        "counts": {
            "input_rows": len(firms) + len(skipped) + len(discarded),
            "passed": len(firms),
            "skipped_no_aum": len(skipped),
            "discarded_below_60pct": len(discarded),
            "high": sum(1 for f in firms if f["priority"] == "High"),
            "medium": sum(1 for f in firms if f["priority"] == "Medium"),
            "low": sum(1 for f in firms if f["priority"] == "Low"),
        },
        "firms": firms,
    }
    out.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    c = payload["counts"]
    print(f"rows in file:        {c['input_rows']}")
    print(f"passed filter:       {c['passed']}  (High {c['high']} / Medium {c['medium']} / Low {c['low']})")
    print(f"skipped (no AUM):    {c['skipped_no_aum']}  {skipped or ''}")
    print(f"discarded (<60%):    {c['discarded_below_60pct']}")
    for d in discarded:
        print(f"  - {d}")
    crd_missing = [f["name"] for f in firms if not f["firm_crd"]]
    print(f"missing CRD:         {len(crd_missing)}  {crd_missing or ''}")
    print(f"wrote:               {out}")


if __name__ == "__main__":
    main()
